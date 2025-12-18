#!/usr/bin/env python3
"""
NullAlleleGenerator - Application Streamlit pour convertir un fichier .gen EASYPOP vers Excel
avec recodage automatique des allèles nuls (10%, 20%, 30%, 40%, 50%, 60%, 70%, 80%)

Développé par: Naffiou KADIRI
Contact: [knaffiou96@gmail.com]

Usage:
    streamlit run null_allele_generator.py
"""

import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import io
from datetime import datetime
import os


def parse_gen_file(file_content):
    """
    Parse le contenu d'un fichier .gen
    Format attendu :
    - Ligne 1: n_pops n_ind max_allele other
    - Lignes 2-21: loc-1, loc-2, ..., loc-20
    - Lignes suivantes: pop / ind allele1 allele2 ...
    """
    lines = file_content.decode('utf-8').strip().split('\n')

    # Ligne 1: métadonnées
    meta = lines[0].split()
    n_pops = int(meta[0])
    n_ind = int(meta[1])
    max_allele = int(meta[2])

    # Lignes 2-21: noms des loci (on les ignore car on utilise loc-1 à loc-20)

    # Trouver où commencent les données
    data_start = 22  # Après "pop"

    # Parser les données
    data_rows = []
    current_pop = 0

    for line in lines[data_start:]:
        line = line.strip()
        if not line or line.lower() == 'pop':
            current_pop += 1
            continue

        parts = line.split(',')
        if len(parts) < 2:
            continue

        # Format: "   1 , 4311 4505 8182 ..."
        ind_part = parts[0].strip()
        alleles_part = parts[1].strip()

        # Extraire les allèles (format: 4 chiffres par paire)
        alleles = alleles_part.split()

        # Convertir chaque paire de 4 chiffres en deux allèles
        row = [current_pop]
        for allele_pair in alleles:
            if len(allele_pair) == 4:
                a1 = int(allele_pair[:2])
                a2 = int(allele_pair[2:])
                row.extend([a1, a2])

        data_rows.append(row)

    return n_pops, n_ind, max_allele, data_rows


def recode_null_alleles(data_rows, threshold):
    """
    Recode les allèles nuls selon un seuil
    """
    recoded = []

    for row in data_rows:
        new_row = [row[0]]  # Population

        # Traiter par paires d'allèles
        for i in range(1, len(row), 2):
            if i + 1 < len(row):
                a1 = row[i]
                a2 = row[i + 1]

                # Règles de recodage
                if a1 <= threshold and a2 <= threshold:
                    # Les deux nuls -> 0, 0
                    new_row.extend([0, 0])
                elif a1 <= threshold:
                    # Seul a1 nul -> a2, a2
                    new_row.extend([a2, a2])
                elif a2 <= threshold:
                    # Seul a2 nul -> a1, a1
                    new_row.extend([a1, a1])
                else:
                    # Aucun nul
                    new_row.extend([a1, a2])

        recoded.append(new_row)

    return recoded


def create_excel_file(n_pops, n_ind, max_allele, data_rows, proportions):
    """
    Crée le fichier Excel avec les sections demandées
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Genotypes"

    n_loci = 20

    # === LIGNE 1 : Métadonnées ===
    ws.cell(1, 1, n_pops)
    ws.cell(1, 2, n_ind)
    ws.cell(1, 3, max_allele)
    ws.cell(1, 4, 3)

    # === LIGNES 2-21 : Noms des loci ===
    for i in range(20):
        ws.cell(2 + i, 1, f"loc-{i + 1}")

    # === LIGNE 22 : Labels des sections ===
    col_offset = 2
    ws.cell(22, col_offset, "0%nulls")
    ws.cell(22, col_offset).font = Font(bold=True)

    sections = {}
    sections[0] = col_offset

    for prop in proportions:
        col_offset += (n_loci * 2)
        ws.cell(22, col_offset, f"{prop}%nulls")
        ws.cell(22, col_offset).font = Font(bold=True)
        sections[prop] = col_offset

    # === LIGNE 23 : En-têtes des colonnes ===
    for prop, start_col in sections.items():
        for locus in range(n_loci):
            ws.cell(23, start_col + (locus * 2), f"loc-{locus + 1}")
            ws.cell(23, start_col + (locus * 2) + 1, f"loc-{locus + 1}")

    # === LIGNES 24+ : Données ===
    # Section 0% (données originales)
    for idx, row in enumerate(data_rows):
        excel_row = 24 + idx
        for col_idx, value in enumerate(row):
            ws.cell(excel_row, col_idx + 1, int(value))

    # Sections recodées
    for prop in proportions:
        threshold = int(prop * max_allele / 100)
        recoded = recode_null_alleles(data_rows, threshold)
        start_col = sections[prop]

        for idx, row in enumerate(recoded):
            excel_row = 24 + idx
            for col_idx in range(1, len(row)):
                ws.cell(excel_row, start_col + col_idx - 1, int(row[col_idx]))

    return wb


def count_zeros(data_rows):
    """Compte le nombre de 0 (données manquantes)"""
    count = 0
    for row in data_rows:
        count += row[1:].count(0)
    return count


def count_homozygotes(data_rows):
    """Compte le nombre d'homozygotes"""
    count = 0
    for row in data_rows:
        for i in range(1, len(row) - 1, 2):
            if row[i] == row[i + 1] and row[i] > 0:
                count += 1
    return count


# === CONFIGURATION DE L'APPLICATION ===
st.set_page_config(
    page_title="Null Allele Generator - Générateur d'Allèles Nuls",
    page_icon="🧬",
    layout="wide",
    menu_items={
        'Get Help': 'mailto:knaffiou96@gmail.com',
        'Report a bug': 'mailto:knaffiou96@gmail.com',
        'About': "## Null Allele Generator\n\nGénérateur de données de génotypes avec allèles nuls simulés pour EASYPOP\n\nDéveloppé par [Votre Nom]"
    }
)

# === EN-TÊTE AVEC INFORMATIONS PERSONNELLES ===
st.title("🧬 Null Allele Generator")
st.markdown("""
### Générateur de Données avec Allèles Nuls Simulés
*Conversion de fichiers .gen EASYPOP vers Excel avec recodage d'allèles nuls*

---
""")

# Section informations personnelles
DEVELOPER_NAME = "Naffiou KADIRI"
DEVELOPER_EMAIL = "knaffiou96@gmail.com"

st.sidebar.markdown("---")

# Sidebar pour les options de recodage
st.sidebar.header("⚙️ Options de Recodage")
st.sidebar.markdown("Sélectionnez les proportions d'allèles nuls à générer :")

# Définir toutes les proportions disponibles
ALL_PROPORTIONS = [10, 20, 30, 40, 50, 60, 70, 80]

# Créer des cases à cocher pour chaque proportion
selected_props = []
col1, col2 = st.sidebar.columns(2)

with col1:
    for prop in ALL_PROPORTIONS[:4]:  # 10%, 20%, 30%, 40%
        if st.checkbox(f"{prop}%", value=(prop in [10, 20, 50])):
            selected_props.append(prop)

with col2:
    for prop in ALL_PROPORTIONS[4:]:  # 50%, 60%, 70%, 80%
        if st.checkbox(f"{prop}%", value=(prop in [10, 20, 50])):
            selected_props.append(prop)

st.sidebar.markdown("---")
st.sidebar.info(
    """
    💡 **Conseil d'utilisation:**
    - La section 0% (original) est toujours incluse
    - Sélectionnez plusieurs proportions pour comparer
    - Vérifiez les statistiques avant téléchargement
    """
)

# Upload du fichier
st.header("📁 1. Chargement du Fichier")
uploaded_file = st.file_uploader(
    "Sélectionnez votre fichier .gen généré par EASYPOP",
    type=['gen'],
    help="Format attendu: fichier de sortie standard d'EASYPOP"
)

if uploaded_file is not None:
    try:
        # Obtenir le nom du fichier d'origine
        original_filename = uploaded_file.name
        base_name = os.path.splitext(original_filename)[0]

        # Afficher le nom du fichier
        st.success(f"**Fichier chargé:** {original_filename}")

        # Lire et parser le fichier
        with st.spinner("Analyse du fichier .gen en cours..."):
            file_content = uploaded_file.read()
            n_pops, n_ind, max_allele, data_rows = parse_gen_file(file_content)

        # Afficher les informations du fichier
        st.header("📊 2. Informations du Fichier")

        # Métriques dans des colonnes
        col1, col2, col3, col4, col5 = st.columns(5)

        with col1:
            st.metric("Populations", n_pops, help="Nombre de sous-populations")
        with col2:
            st.metric("Ind/Pop", n_ind, help="Individus par population")
        with col3:
            st.metric("Total Ind", len(data_rows), help="Nombre total d'individus")
        with col4:
            st.metric("Allèle Max", max_allele, help="Valeur maximale des allèles")
        with col5:
            loci_count = (len(data_rows[0]) - 1) // 2 if data_rows else 0
            st.metric("Loci", loci_count, help="Nombre de loci génétiques")

        # Aperçu des données
        with st.expander("🔍 Aperçu des Données Brutes"):
            preview_df = pd.DataFrame(data_rows[:10])
            preview_df.columns = ["Pop"] + [f"Loc-{i // 2 + 1}_{'A' if i % 2 == 0 else 'B'}"
                                            for i in range(len(preview_df.columns) - 1)]
            st.dataframe(preview_df, use_container_width=True)

            # Statistiques de base
            st.subheader("Statistiques Préliminaires")
            total_alleles = sum(len(row) - 1 for row in data_rows)
            missing_original = count_zeros(data_rows)
            homozygotes_original = count_homozygotes(data_rows)

            col_stat1, col_stat2, col_stat3 = st.columns(3)
            with col_stat1:
                st.metric("Allèles totaux", total_alleles)
            with col_stat2:
                st.metric("Données manquantes (0%)", missing_original)
            with col_stat3:
                st.metric("Homozygotes (0%)", homozygotes_original)

        # Vérifier qu'au moins une proportion est sélectionnée
        if not selected_props:
            st.warning("""
            ⚠️ **Aucune proportion sélectionnée**

            Veuillez sélectionner au moins une proportion d'allèles nuls dans la barre latérale.
            Le fichier généré contiendra uniquement les données originales (0%).
            """)

            # Demander confirmation pour générer seulement 0%
            if st.button("Générer uniquement les données originales (0%)", type="secondary"):
                selected_props = []
                generate_clicked = True
            else:
                generate_clicked = False
        else:
            st.success(f"✅ **Proportions sélectionnées:** {', '.join([f'{p}%' for p in sorted(selected_props)])}")
            generate_clicked = st.button("🔄 Générer le Fichier Excel", type="primary")

        if (not selected_props and 'generate_clicked' in locals() and generate_clicked) or \
                (selected_props and generate_clicked):

            with st.spinner("Génération du fichier Excel en cours..."):
                # Créer le fichier Excel
                wb = create_excel_file(n_pops, n_ind, max_allele, data_rows, sorted(selected_props))

                # Calculer les statistiques
                st.header("📈 3. Statistiques des Données Recodées")

                # Stats pour chaque section
                stats_data = []

                # 0% (original)
                zeros_0 = count_zeros(data_rows)
                homo_0 = count_homozygotes(data_rows)
                stats_data.append(["0% (Original)", zeros_0, homo_0,
                                   f"{(zeros_0 / (total_alleles if total_alleles > 0 else 1) * 100):.1f}%"])

                # Sections recodées
                for prop in sorted(selected_props):
                    threshold = int(prop * max_allele / 100)
                    recoded = recode_null_alleles(data_rows, threshold)
                    zeros = count_zeros(recoded)
                    homo = count_homozygotes(recoded)
                    missing_percent = (zeros / (total_alleles if total_alleles > 0 else 1) * 100)
                    stats_data.append([f"{prop}% (Seuil: ≤{threshold})", zeros, homo, f"{missing_percent:.1f}%"])

                stats_df = pd.DataFrame(
                    stats_data,
                    columns=["Section (Seuil)", "Données manquantes (0)", "Homozygotes", "% Manquants"]
                )

                # Afficher le tableau
                st.dataframe(stats_df, use_container_width=True)

                # Visualisations
                st.subheader("📊 Visualisations")
                viz_col1, viz_col2 = st.columns(2)

                with viz_col1:
                    chart_data = stats_df.copy()
                    chart_data['Section'] = chart_data['Section (Seuil)'].str.split(' ').str[0]
                    st.bar_chart(
                        chart_data.set_index("Section")["Données manquantes (0)"],
                        use_container_width=True
                    )
                    st.caption("**Évolution des données manquantes**")

                with viz_col2:
                    st.bar_chart(
                        chart_data.set_index("Section")["Homozygotes"],
                        use_container_width=True
                    )
                    st.caption("**Évolution des homozygotes**")

                # Sauvegarder en mémoire
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)

                st.success("✅ **Fichier Excel généré avec succès !**")

                # Bouton de téléchargement
                st.header("💾 4. Téléchargement")

                # Utiliser le nom du fichier d'origine pour le fichier de sortie
                filename = f"genotypes_recoded_{base_name}.xlsx"

                col_dl1, col_dl2 = st.columns([2, 1])

                with col_dl1:
                    st.download_button(
                        label=f"📥 Télécharger '{filename}'",
                        data=excel_buffer,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True
                    )

                with col_dl2:
                    # Option pour renommer le fichier
                    custom_name = st.text_input("Nom personnalisé", value=base_name)
                    if custom_name != base_name:
                        alt_filename = f"genotypes_recoded_{custom_name}.xlsx"
                        st.download_button(
                            label="📥 Télécharger avec nom personnalisé",
                            data=excel_buffer,
                            file_name=alt_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )

                st.info(f"""
                📂 **Fichier généré:** `{filename}`

                **Contenu:**
                - Données originales (section 0%)
                - {len(selected_props)} section(s) avec allèles nuls recodés
                - Format Excel compatible avec la plupart des logiciels
                """)

    except Exception as e:
        st.error(f"❌ **Erreur lors du traitement**")
        st.code(str(e), language='text')

        with st.expander("🔧 Dépannage"):
            st.markdown("""
            **Problèmes courants et solutions:**

            1. **Format de fichier incorrect**
               - Vérifiez que le fichier est bien un .gen généré par EASYPOP
               - Vérifiez le format des données (lignes 2-21 doivent contenir les noms des loci)

            2. **Erreur de décodage**
               - Assurez-vous que le fichier est encodé en UTF-8
               - Essayez de réenregistrer le fichier depuis EASYPOP

            3. **Structure des données**
               - Le fichier doit commencer par: `n_pops n_ind max_allele other`
               - Les données doivent suivre le format standard EASYPOP

            **Besoin d'aide?** Contactez: {DEVELOPER_EMAIL}
            """.format(DEVELOPER_EMAIL=DEVELOPER_EMAIL))

else:
    # Écran d'accueil
    st.info("👆 **Bienvenue dans Null Allele Generator!** Veuillez charger un fichier .gen pour commencer.")

    # Guide d'utilisation
    with st.expander("🎯 **Guide Complet d'Utilisation**", expanded=True):
        st.markdown(f"""
        ### 🧬 **Null Allele Generator v1.0.0**

        ---

        #### **📋 Fonctionnalités Principales**

        1. **Conversion de format**  
           - Fichiers .gen (EASYPOP) → Excel (.xlsx)
           - Structure préservée avec métadonnées

        2. **Simulation d'allèles nuls**  
           - 8 proportions disponibles: 10%, 20%, 30%, 40%, 50%, 60%, 70%, 80%
           - Recodage automatique selon des règles définies
           - Comparaison multiple en un seul fichier

        3. **Analyse statistique**  
           - Comptage des données manquantes
           - Détection des homozygotes
           - Visualisations intégrées

        4. **Personnalisation**  
           - Sélection des proportions souhaitées
           - Nom personnalisé pour le fichier de sortie
           - Interface configurable

        ---

        #### **⚙️ Règles de Recodage**

        Pour chaque proportion P% (seuil = P% × max_allele):

        | Condition | Résultat | Signification |
        |-----------|----------|---------------|
        | Allèle1 ≤ seuil ET Allèle2 ≤ seuil | 0, 0 | Donnée manquante |
        | Allèle1 ≤ seuil ET Allèle2 > seuil | Allèle2, Allèle2 | Homozygote pour allèle2 |
        | Allèle1 > seuil ET Allèle2 ≤ seuil | Allèle1, Allèle1 | Homozygote pour allèle1 |
        | Allèle1 > seuil ET Allèle2 > seuil | Allèle1, Allèle2 | Donnée inchangée |

        ---

        #### **📁 Format de Sortie Excel**

        Le fichier généré contient:

        1. **Ligne 1:** Métadonnées (`n_pops n_ind max_allele 3`)
        2. **Lignes 2-21:** Noms des loci (`loc-1` à `loc-20`)
        3. **Ligne 22:** Sections (`0%nulls`, `10%nulls`, etc.)
        4. **Ligne 23:** En-têtes des colonnes
        5. **Lignes 24+:** Données des génotypes

        ---

        #### **🔧 Prérequis Techniques**

        - Fichier .gen généré par **EASYPOP**
        - Format standard (20 loci)
        - Encodage UTF-8
        - Navigateur web moderne
        """)

# Pied de page avec crédits
st.markdown("---")
footer_col1, footer_col2, footer_col3 = st.columns([2, 1, 1])

with footer_col1:
    st.caption(f"🧬 **Null Allele Generator** v1.0.0 | Développé par {DEVELOPER_NAME}")

with footer_col2:
    st.caption(f"📧 [{DEVELOPER_EMAIL}](mailto:{DEVELOPER_EMAIL})")

with footer_col3:
    st.caption(f"📅 {datetime.now().strftime('%d/%m/%Y %H:%M')}")